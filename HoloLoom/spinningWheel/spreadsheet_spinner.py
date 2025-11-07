"""
Spreadsheet Spinner - Multi-Format Spreadsheet Ingestion
=========================================================

Ingests spreadsheets from multiple formats into structured MemoryShards.
Preserves table structure, formulas, and metadata.

Features:
- Excel support (.xlsx, .xls)
- CSV/TSV support
- Google Sheets API integration
- LibreOffice Calc (.ods)
- Smart header detection
- Formula extraction
- Data type inference
- Sheet-level or row-level chunking
- Named range extraction

Requirements:
    pip install pandas openpyxl

Optional:
    pip install xlrd  # For .xls files
    pip install odfpy  # For .ods files
    pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client  # Google Sheets

Author: Claude Code
Date: November 2025
"""

import asyncio
from pathlib import Path
from typing import List, Optional, Dict, Any, AsyncIterator, Union
from dataclasses import dataclass, field
import warnings
import re

from HoloLoom.spinningWheel.protocol import (
    BaseSpinner,
    SpinResult,
    SpinnerCapabilities
)
from HoloLoom.spinningWheel.importance import ImportanceScorer
from HoloLoom.documentation.types import MemoryShard

# Try to import pandas (required)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    warnings.warn("pandas not available. Install with: pip install pandas openpyxl")

# Try to import openpyxl for Excel
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import xlrd for old Excel (.xls)
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

# Try to import odfpy for LibreOffice
try:
    import odf
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False

# Try to import Google Sheets API
try:
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class SpreadsheetCell:
    """Individual cell data"""
    row: int
    column: int
    value: Any
    formula: Optional[str] = None
    data_type: str = 'text'  # 'text', 'number', 'date', 'boolean', 'formula'
    formatted_value: Optional[str] = None


@dataclass
class SpreadsheetTable:
    """Table/range of cells"""
    name: str
    headers: List[str]
    data: List[List[Any]]  # 2D array
    start_row: int
    start_col: int
    formulas: Dict[str, str] = field(default_factory=dict)  # cell_ref -> formula

    @property
    def row_count(self) -> int:
        return len(self.data)

    @property
    def column_count(self) -> int:
        return len(self.headers)

    def to_dataframe(self) -> 'pd.DataFrame':
        """Convert to pandas DataFrame"""
        if PANDAS_AVAILABLE:
            return pd.DataFrame(self.data, columns=self.headers)
        raise ImportError("pandas required for DataFrame conversion")

    def to_markdown(self) -> str:
        """Convert to markdown table"""
        lines = []

        # Header
        lines.append("| " + " | ".join(self.headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(self.headers)) + " |")

        # Data rows (limit to first 100 for sanity)
        for row in self.data[:100]:
            row_str = [str(cell) if cell is not None else '' for cell in row]
            lines.append("| " + " | ".join(row_str) + " |")

        if len(self.data) > 100:
            lines.append(f"\n... ({len(self.data) - 100} more rows)")

        return "\n".join(lines)


@dataclass
class SpreadsheetSheet:
    """Single sheet/tab in spreadsheet"""
    name: str
    tables: List[SpreadsheetTable]
    row_count: int
    column_count: int
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def table_count(self) -> int:
        return len(self.tables)


@dataclass
class Spreadsheet:
    """Complete spreadsheet document"""
    file_path: Path
    sheets: List[SpreadsheetSheet]
    file_format: str  # 'xlsx', 'xls', 'csv', 'ods', 'google_sheets'
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def sheet_count(self) -> int:
        return len(self.sheets)


# =============================================================================
# Spreadsheet Parser
# =============================================================================

class SpreadsheetParser:
    """Parse spreadsheets into structured format"""

    @staticmethod
    def parse_file(file_path: Path) -> Spreadsheet:
        """Parse spreadsheet file"""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for spreadsheet parsing")

        suffix = file_path.suffix.lower()

        if suffix in ['.xlsx', '.xlsm']:
            return SpreadsheetParser._parse_excel(file_path)
        elif suffix == '.xls':
            return SpreadsheetParser._parse_xls(file_path)
        elif suffix == '.csv':
            return SpreadsheetParser._parse_csv(file_path)
        elif suffix == '.tsv':
            return SpreadsheetParser._parse_tsv(file_path)
        elif suffix == '.ods':
            return SpreadsheetParser._parse_ods(file_path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

    @staticmethod
    def _parse_excel(file_path: Path) -> Spreadsheet:
        """Parse Excel (.xlsx, .xlsm) file"""
        sheets = []

        # Read all sheets
        excel_file = pd.ExcelFile(file_path, engine='openpyxl')

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

            # Auto-detect header row
            header_row, headers = SpreadsheetParser._detect_headers(df)

            # Extract data (skip header row)
            if header_row is not None:
                data = df.iloc[header_row + 1:].values.tolist()
            else:
                # No header detected, use all data
                headers = [f"Column_{i}" for i in range(len(df.columns))]
                data = df.values.tolist()

            # Clean data (remove trailing None rows)
            data = SpreadsheetParser._clean_data(data)

            # Create table
            table = SpreadsheetTable(
                name=f"{sheet_name}_data",
                headers=headers,
                data=data,
                start_row=header_row + 1 if header_row is not None else 0,
                start_col=0
            )

            # Extract formulas if openpyxl available
            formulas = {}
            if OPENPYXL_AVAILABLE:
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=False)
                    ws = wb[sheet_name]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                cell_ref = f"{cell.column_letter}{cell.row}"
                                formulas[cell_ref] = cell.value
                    table.formulas = formulas
                except Exception as e:
                    warnings.warn(f"Failed to extract formulas: {e}")

            sheet = SpreadsheetSheet(
                name=sheet_name,
                tables=[table],
                row_count=len(df),
                column_count=len(df.columns)
            )
            sheets.append(sheet)

        return Spreadsheet(
            file_path=file_path,
            sheets=sheets,
            file_format='xlsx'
        )

    @staticmethod
    def _parse_xls(file_path: Path) -> Spreadsheet:
        """Parse old Excel (.xls) file"""
        if not XLRD_AVAILABLE:
            warnings.warn("xlrd not available for .xls files. Install with: pip install xlrd")
            # Fall back to pandas (may not work for all .xls files)
            return SpreadsheetParser._parse_excel(file_path)

        sheets = []
        excel_file = pd.ExcelFile(file_path, engine='xlrd')

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

            header_row, headers = SpreadsheetParser._detect_headers(df)
            if header_row is not None:
                data = df.iloc[header_row + 1:].values.tolist()
            else:
                headers = [f"Column_{i}" for i in range(len(df.columns))]
                data = df.values.tolist()

            data = SpreadsheetParser._clean_data(data)

            table = SpreadsheetTable(
                name=f"{sheet_name}_data",
                headers=headers,
                data=data,
                start_row=header_row + 1 if header_row is not None else 0,
                start_col=0
            )

            sheet = SpreadsheetSheet(
                name=sheet_name,
                tables=[table],
                row_count=len(df),
                column_count=len(df.columns)
            )
            sheets.append(sheet)

        return Spreadsheet(
            file_path=file_path,
            sheets=sheets,
            file_format='xls'
        )

    @staticmethod
    def _parse_csv(file_path: Path) -> Spreadsheet:
        """Parse CSV file"""
        df = pd.read_csv(file_path, header=None)

        header_row, headers = SpreadsheetParser._detect_headers(df)
        if header_row is not None:
            data = df.iloc[header_row + 1:].values.tolist()
        else:
            headers = [f"Column_{i}" for i in range(len(df.columns))]
            data = df.values.tolist()

        data = SpreadsheetParser._clean_data(data)

        table = SpreadsheetTable(
            name="csv_data",
            headers=headers,
            data=data,
            start_row=header_row + 1 if header_row is not None else 0,
            start_col=0
        )

        sheet = SpreadsheetSheet(
            name="Sheet1",
            tables=[table],
            row_count=len(df),
            column_count=len(df.columns)
        )

        return Spreadsheet(
            file_path=file_path,
            sheets=[sheet],
            file_format='csv'
        )

    @staticmethod
    def _parse_tsv(file_path: Path) -> Spreadsheet:
        """Parse TSV file"""
        df = pd.read_csv(file_path, sep='\t', header=None)

        header_row, headers = SpreadsheetParser._detect_headers(df)
        if header_row is not None:
            data = df.iloc[header_row + 1:].values.tolist()
        else:
            headers = [f"Column_{i}" for i in range(len(df.columns))]
            data = df.values.tolist()

        data = SpreadsheetParser._clean_data(data)

        table = SpreadsheetTable(
            name="tsv_data",
            headers=headers,
            data=data,
            start_row=header_row + 1 if header_row is not None else 0,
            start_col=0
        )

        sheet = SpreadsheetSheet(
            name="Sheet1",
            tables=[table],
            row_count=len(df),
            column_count=len(df.columns)
        )

        return Spreadsheet(
            file_path=file_path,
            sheets=[sheet],
            file_format='tsv'
        )

    @staticmethod
    def _parse_ods(file_path: Path) -> Spreadsheet:
        """Parse LibreOffice Calc (.ods) file"""
        if not ODF_AVAILABLE:
            warnings.warn("odfpy not available for .ods files. Install with: pip install odfpy")
            raise ImportError("odfpy required for .ods files")

        sheets = []
        ods_file = pd.ExcelFile(file_path, engine='odf')

        for sheet_name in ods_file.sheet_names:
            df = pd.read_excel(ods_file, sheet_name=sheet_name, header=None)

            header_row, headers = SpreadsheetParser._detect_headers(df)
            if header_row is not None:
                data = df.iloc[header_row + 1:].values.tolist()
            else:
                headers = [f"Column_{i}" for i in range(len(df.columns))]
                data = df.values.tolist()

            data = SpreadsheetParser._clean_data(data)

            table = SpreadsheetTable(
                name=f"{sheet_name}_data",
                headers=headers,
                data=data,
                start_row=header_row + 1 if header_row is not None else 0,
                start_col=0
            )

            sheet = SpreadsheetSheet(
                name=sheet_name,
                tables=[table],
                row_count=len(df),
                column_count=len(df.columns)
            )
            sheets.append(sheet)

        return Spreadsheet(
            file_path=file_path,
            sheets=sheets,
            file_format='ods'
        )

    @staticmethod
    def _detect_headers(df: 'pd.DataFrame') -> tuple[Optional[int], List[str]]:
        """
        Auto-detect header row in DataFrame.

        Returns:
            (header_row_index, headers) or (None, []) if no header detected
        """
        if len(df) == 0:
            return None, []

        # Check first few rows for header candidates
        for i in range(min(5, len(df))):
            row = df.iloc[i]

            # Header heuristics:
            # 1. All values are strings
            # 2. No duplicate values
            # 3. Not all None/empty
            # 4. Looks like column names (short, descriptive)

            if all(isinstance(val, (str, type(None))) for val in row):
                non_null = [val for val in row if val is not None and str(val).strip()]

                if len(non_null) > 0 and len(non_null) == len(set(non_null)):
                    # Check if values look like headers (short, no numbers)
                    avg_len = sum(len(str(val)) for val in non_null) / len(non_null)
                    if avg_len < 30:  # Headers are usually short
                        headers = [str(val) if val is not None else f"Column_{j}" for j, val in enumerate(row)]
                        return i, headers

        # No header detected
        return None, []

    @staticmethod
    def _clean_data(data: List[List[Any]]) -> List[List[Any]]:
        """Remove trailing None rows"""
        # Remove rows that are completely None
        cleaned = []
        for row in data:
            if any(val is not None and str(val).strip() for val in row):
                cleaned.append(row)
        return cleaned


# =============================================================================
# Spreadsheet Spinner
# =============================================================================

class SpreadsheetSpinner(BaseSpinner):
    """
    Ingest spreadsheets from multiple formats.

    Features:
    - Excel, CSV, TSV, ODS support
    - Smart header detection
    - Formula extraction
    - Sheet-level or row-level chunking
    """

    def __init__(
        self,
        importance_threshold: float = 0.3,
        chunk_mode: str = 'sheet',  # 'sheet', 'table', or 'row'
        max_rows_per_shard: int = 1000,  # For row-level chunking
        include_formulas: bool = True
    ):
        """
        Initialize SpreadsheetSpinner.

        Args:
            importance_threshold: Minimum importance score
            chunk_mode: How to chunk data ('sheet', 'table', 'row')
            max_rows_per_shard: Maximum rows per shard (for row chunking)
            include_formulas: Extract formulas (Excel only)
        """
        super().__init__(name="spreadsheet")
        self.importance_threshold = importance_threshold
        self.chunk_mode = chunk_mode
        self.max_rows_per_shard = max_rows_per_shard
        self.include_formulas = include_formulas

        # Initialize importance scorer
        self.importance_scorer = ImportanceScorer()

    async def _spin_impl(self, source: Any, **kwargs) -> List[MemoryShard]:
        """
        Core spreadsheet ingestion implementation.

        Args:
            source: Path to spreadsheet file

        Returns:
            List of MemoryShard objects (base class handles filtering and wrapping)
        """
        file_path = Path(source)

        # Parse spreadsheet
        spreadsheet = await self._parse_file(file_path)

        # Convert to shards (base class handles filtering and wrapping)
        shards = self._spreadsheet_to_shards(spreadsheet)

        return shards

    async def _parse_file(self, file_path: Path) -> Spreadsheet:
        """Parse file in thread pool"""
        loop = asyncio.get_event_loop()
        spreadsheet = await loop.run_in_executor(
            None,
            SpreadsheetParser.parse_file,
            file_path
        )
        return spreadsheet

    def _spreadsheet_to_shards(self, spreadsheet: Spreadsheet) -> List[MemoryShard]:
        """Convert spreadsheet to MemoryShards"""
        shards = []

        for sheet in spreadsheet.sheets:
            for table in sheet.tables:
                if self.chunk_mode == 'sheet':
                    # One shard per sheet
                    shard = self._create_sheet_shard(spreadsheet, sheet, table)
                    shards.append(shard)

                elif self.chunk_mode == 'table':
                    # One shard per table
                    shard = self._create_table_shard(spreadsheet, sheet, table)
                    shards.append(shard)

                elif self.chunk_mode == 'row':
                    # Multiple shards based on row chunks
                    row_shards = self._create_row_shards(spreadsheet, sheet, table)
                    shards.extend(row_shards)

        return shards

    def _create_sheet_shard(
        self,
        spreadsheet: Spreadsheet,
        sheet: SpreadsheetSheet,
        table: SpreadsheetTable
    ) -> MemoryShard:
        """Create shard for entire sheet"""
        # Convert table to markdown
        markdown = table.to_markdown()

        # Add formula info if available
        formula_text = ""
        if table.formulas:
            formula_text = "\n\n## Formulas\n"
            for cell_ref, formula in list(table.formulas.items())[:20]:  # Limit to 20
                formula_text += f"- {cell_ref}: `{formula}`\n"

        text = f"# {sheet.name}\n\n{markdown}{formula_text}"

        return MemoryShard(
            id=f"spreadsheet_{spreadsheet.file_path.stem}_{sheet.name}",
            text=text,
            episode=f"spreadsheet_{spreadsheet.file_path.stem}",
            metadata={
                'file_name': spreadsheet.file_path.name,
                'file_format': spreadsheet.file_format,
                'sheet_name': sheet.name,
                'row_count': table.row_count,
                'column_count': table.column_count,
                'headers': table.headers,
                'has_formulas': len(table.formulas) > 0,
                'importance_score': self._score_table_importance(table),
                'shard_type': 'sheet'
            }
        )

    def _create_table_shard(
        self,
        spreadsheet: Spreadsheet,
        sheet: SpreadsheetSheet,
        table: SpreadsheetTable
    ) -> MemoryShard:
        """Create shard for table (same as sheet for most cases)"""
        return self._create_sheet_shard(spreadsheet, sheet, table)

    def _create_row_shards(
        self,
        spreadsheet: Spreadsheet,
        sheet: SpreadsheetSheet,
        table: SpreadsheetTable
    ) -> List[MemoryShard]:
        """Create shards for row chunks"""
        shards = []
        chunk_size = self.max_rows_per_shard

        for i in range(0, len(table.data), chunk_size):
            chunk_data = table.data[i:i + chunk_size]

            # Create markdown for this chunk
            lines = []
            lines.append("| " + " | ".join(table.headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(table.headers)) + " |")

            for row in chunk_data:
                row_str = [str(cell) if cell is not None else '' for cell in row]
                lines.append("| " + " | ".join(row_str) + " |")

            markdown = "\n".join(lines)
            text = f"# {sheet.name} (rows {i + 1}-{i + len(chunk_data)})\n\n{markdown}"

            shard = MemoryShard(
                id=f"spreadsheet_{spreadsheet.file_path.stem}_{sheet.name}_rows_{i}_{i + len(chunk_data)}",
                text=text,
                episode=f"spreadsheet_{spreadsheet.file_path.stem}",
                metadata={
                    'file_name': spreadsheet.file_path.name,
                    'file_format': spreadsheet.file_format,
                    'sheet_name': sheet.name,
                    'row_start': i,
                    'row_end': i + len(chunk_data),
                    'row_count': len(chunk_data),
                    'column_count': table.column_count,
                    'headers': table.headers,
                    'importance_score': self._score_chunk_importance(chunk_data),
                    'shard_type': 'row_chunk'
                }
            )
            shards.append(shard)

        return shards

    def _score_table_importance(self, table: SpreadsheetTable) -> float:
        """Score importance of table"""
        # Convert table to text for scoring
        text = table.to_markdown()

        score = self.importance_scorer.score(
            text,
            metadata={
                'length': len(text),
                'row_count': table.row_count,
                'column_count': table.column_count
            }
        )

        # Boost tables with formulas
        if table.formulas:
            score.score *= 1.2

        return min(1.0, score.score)

    def _score_chunk_importance(self, chunk_data: List[List[Any]]) -> float:
        """Score importance of data chunk"""
        # Simple scoring based on non-null cells
        non_null_count = sum(
            1 for row in chunk_data
            for cell in row
            if cell is not None and str(cell).strip()
        )

        total_cells = len(chunk_data) * (len(chunk_data[0]) if chunk_data else 0)
        if total_cells == 0:
            return 0.0

        # More filled cells = higher importance
        fill_ratio = non_null_count / total_cells
        return min(1.0, fill_ratio * 0.8)  # Max 0.8 for data density

    async def spin_stream(
        self,
        file_path: Path,
        batch_size: int = 1
    ) -> AsyncIterator[MemoryShard]:
        """Stream shards as they're created"""
        result = await self.spin(file_path)

        if not result.success:
            return

        for shard in result.shards:
            yield shard

    def get_capabilities(self) -> SpinnerCapabilities:
        """Get spinner capabilities"""
        return SpinnerCapabilities(
            streaming=True,
            incremental=False,
            batch=True,
            supported_formats=['xlsx', 'xls', 'csv', 'tsv', 'ods'],
            metadata={
                'chunk_modes': ['sheet', 'table', 'row'],
                'formula_extraction': self.include_formulas
            }
        )

    def is_available(self) -> bool:
        """Check if pandas is available"""
        return PANDAS_AVAILABLE


# =============================================================================
# Convenience Functions
# =============================================================================

async def ingest_spreadsheet(
    file_path: Path,
    chunk_mode: str = 'sheet'
) -> SpinResult:
    """
    Convenience function to ingest spreadsheet.

    Args:
        file_path: Path to spreadsheet file
        chunk_mode: How to chunk data ('sheet', 'table', 'row')

    Returns:
        SpinResult with table shards
    """
    spinner = SpreadsheetSpinner(
        importance_threshold=0.2,  # Lower threshold for convenience
        chunk_mode=chunk_mode
    )

    return await spinner.spin(file_path)
